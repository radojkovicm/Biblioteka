import os
from datetime import datetime
from typing import Optional
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models.book import Book
from app.models.book_copy import BookCopy
from app.models.member import Member

EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "exports")


def _ensure_export_dir():
    os.makedirs(EXPORT_DIR, exist_ok=True)


def export_books_to_excel(db: Session = None) -> str:
    _ensure_export_dir()
    close_db = False
    if db is None:
        db = SessionLocal()
        close_db = True

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Knjige"
        ws.append(["Inventarni broj", "Naslov", "Autor", "Izdavač", "Godina", "Žanr", "Jezik", "Polica", "Status", "Stanje"])

        copies = db.query(BookCopy).filter(BookCopy.is_deleted == False).all()
        for copy in copies:
            book = db.query(Book).filter(Book.id == copy.book_id).first()
            ws.append([
                copy.library_number,
                book.title if book else "",
                book.author if book else "",
                book.publisher if book else "",
                book.year_published if book else "",
                book.genre if book else "",
                book.language if book else "",
                copy.shelf_location or "",
                copy.status,
                copy.condition,
            ])

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        path = os.path.join(EXPORT_DIR, f"knjige_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(path)
        return path
    finally:
        if close_db:
            db.close()


def export_members_to_excel(db: Session = None) -> str:
    _ensure_export_dir()
    close_db = False
    if db is None:
        db = SessionLocal()
        close_db = True

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Članovi"
        ws.append(["Broj člana", "Ime", "Prezime", "Datum rođenja", "Email", "Telefon", "Adresa", "Tip", "Aktivan", "Blokiran"])

        members = db.query(Member).filter(Member.is_deleted == False).all()
        for m in members:
            ws.append([
                m.member_number,
                m.first_name,
                m.last_name,
                str(m.date_of_birth) if m.date_of_birth else "",
                m.email or "",
                m.phone or "",
                m.address or "",
                m.member_type,
                "Da" if m.is_active else "Ne",
                "Da" if m.is_blocked else "Ne",
            ])

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        path = os.path.join(EXPORT_DIR, f"clanovi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(path)
        return path
    finally:
        if close_db:
            db.close()


def generate_import_template(template_type: str) -> str:
    """Generate empty Excel template for import."""
    _ensure_export_dir()
    wb = Workbook()
    ws = wb.active

    if template_type == "books":
        ws.title = "Šablon - Knjige"
        ws.append(["library_number", "title", "author", "publisher", "year_published", "genre", "language", "shelf_location"])
        ws.append(["INV-00001", "Primer naslova", "Ime Autora", "Izdavač", 2020, "roman", "srpski", "A-01"])
    elif template_type == "members":
        ws.title = "Šablon - Članovi"
        ws.append(["member_number", "first_name", "last_name", "date_of_birth", "email", "phone", "address", "member_type"])
        ws.append(["MBR-00001", "Petar", "Petrović", "1990-01-15", "petar@example.com", "0601234567", "Adresa 1", "odrasli"])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4

    path = os.path.join(EXPORT_DIR, f"sablon_{template_type}.xlsx")
    wb.save(path)
    return path


def import_books_from_excel(file_path: str, db: Session) -> dict:
    """Import books from Excel file. Returns stats."""
    wb = load_workbook(file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    required = {"library_number", "title", "author"}
    if not required.issubset(set(headers)):
        return {"success": False, "error": f"Nedostaju kolone: {required - set(headers)}", "imported": 0, "errors": []}

    imported = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))
        if not data.get("library_number") or not data.get("title") or not data.get("author"):
            errors.append(f"Red {row_num}: nedostaju obavezna polja")
            continue

        # Check duplicate library_number
        existing = db.query(BookCopy).filter(BookCopy.library_number == str(data["library_number"])).first()
        if existing:
            errors.append(f"Red {row_num}: inventarni broj {data['library_number']} već postoji")
            continue

        # Find or create book
        book = db.query(Book).filter(
            Book.title == str(data["title"]),
            Book.author == str(data["author"]),
            Book.is_deleted == False,
        ).first()

        if not book:
            book = Book(
                title=str(data["title"]),
                author=str(data["author"]),
                publisher=str(data.get("publisher") or ""),
                year_published=int(data["year_published"]) if data.get("year_published") else None,
                genre=str(data.get("genre") or ""),
                language=str(data.get("language") or "srpski"),
            )
            db.add(book)
            db.flush()

        copy = BookCopy(
            library_number=str(data["library_number"]),
            book_id=book.id,
            shelf_location=str(data.get("shelf_location") or ""),
            status="available",
        )
        db.add(copy)
        book.total_copies = db.query(BookCopy).filter(
            BookCopy.book_id == book.id, BookCopy.is_deleted == False
        ).count() + 1
        imported += 1

    db.commit()
    return {"success": True, "imported": imported, "errors": errors}


def import_members_from_excel(file_path: str, db: Session) -> dict:
    """Import members from Excel file."""
    wb = load_workbook(file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    required = {"first_name", "last_name"}
    if not required.issubset(set(headers)):
        return {"success": False, "error": f"Nedostaju kolone: {required - set(headers)}", "imported": 0, "errors": []}

    imported = 0
    errors = []
    valid_types = {"djak", "student", "odrasli", "penzioner", "institucija"}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))
        if not data.get("first_name") or not data.get("last_name"):
            errors.append(f"Red {row_num}: nedostaju obavezna polja")
            continue

        member_type = str(data.get("member_type") or "odrasli").lower()
        if member_type not in valid_types:
            errors.append(f"Red {row_num}: nepoznat tip člana '{member_type}'")
            continue

        # Generate member number if not provided
        member_number = data.get("member_number")
        if member_number:
            existing = db.query(Member).filter(Member.member_number == str(member_number)).first()
            if existing:
                errors.append(f"Red {row_num}: broj člana {member_number} već postoji")
                continue
        else:
            last = db.query(Member).order_by(Member.id.desc()).first()
            next_num = (last.id + 1) if last else 1
            member_number = f"MBR-{next_num:05d}"

        dob = None
        if data.get("date_of_birth"):
            try:
                dob_val = data["date_of_birth"]
                if isinstance(dob_val, str):
                    from datetime import datetime as dt
                    dob = dt.strptime(dob_val, "%Y-%m-%d").date()
                else:
                    dob = dob_val
            except (ValueError, TypeError):
                pass

        member = Member(
            member_number=str(member_number),
            first_name=str(data["first_name"]),
            last_name=str(data["last_name"]),
            date_of_birth=dob,
            email=str(data.get("email") or "") or None,
            phone=str(data.get("phone") or "") or None,
            address=str(data.get("address") or "") or None,
            member_type=member_type,
        )
        db.add(member)
        imported += 1

    db.commit()
    return {"success": True, "imported": imported, "errors": errors}
